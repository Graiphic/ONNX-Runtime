# report.py

import os
import re
import subprocess
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.chart import Reference
from openpyxl.chart.pie_chart import PieChart
from openpyxl.chart.series import DataPoint, GraphicalProperties
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.colors import ColorChoice
from openpyxl.utils import get_column_letter
from openpyxl.chart.layout import Layout, ManualLayout
from collections import Counter
import onnx
import onnxruntime as ort
import platform
import urllib.parse
from collections import defaultdict



try:
    import cpuinfo  # To get a more readable CPU name
except ImportError:
    cpuinfo = None

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt


# --------------------------------------------------
# Excel Report Generation: Detailed (non-aggregated)
# --------------------------------------------------

def generate_report(results, provider, profiling_dir, models_dir):
    """
    Generates an Excel file report_{provider}.xlsx from the `results` list.
    Status categories handled:
      - SUCCESS
      - SUCCESS (via decomposition)
      - SUCCESS WITH FALLBACK
      - SUCCESS WITH FALLBACK (via decomposition)
      - UNKNOWN (no Node event)
      - NOT TESTED
      - SKIPPED
      - FAIL
    """

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(["op", "provider", "used_provider", "status"])

    color_map = {
        "SUCCESS": "00AA44",
        "SUCCESS (via decomposition)": "007700",
        "SUCCESS WITH FALLBACK": "FFAA00",
        "SUCCESS WITH FALLBACK (via decomposition)": "FF7700",
        "UNKNOWN (no Node event)": "DEDEDE",
        "NOT TESTED": "4D7CFE",
        "SKIPPED": "CCCCCC",
        "FAIL": "FF0000",
    }

    def get_fill_for_status(status):
        for key in color_map:
            if status == key or status.startswith(key):
                return PatternFill("solid", fgColor=color_map[key])
        return PatternFill("solid", fgColor=color_map["FAIL"])

    for row in results:
        op_name, provider_used, used_provider, status = row
        display_status = "SUCCESS" if status.startswith("SUCCESS (with complexification)") else status
        ws.append([op_name, provider_used, used_provider or "", display_status])
        ws.cell(row=ws.max_row, column=4).fill = get_fill_for_status(display_status)

    for col_idx in range(1, 5):
        max_len = max(len(str(cell.value)) for cell in ws[get_column_letter(col_idx)] if cell.value)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)

    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = border

    counter = Counter()
    for _, _, _, status in results:
        normalized = "SUCCESS" if status.startswith("SUCCESS (with complexification)") else status
        matched = False
        for key in color_map:
            if normalized == key or normalized.startswith(key):
                counter[key] += 1
                matched = True
                break
        if not matched:
            counter["FAIL"] += 1

    ws_data = wb.create_sheet("Data_PieChart")
    ws_data.append(["Category", "Count", "Percentage"])
    total = sum(counter.values()) or 1

    ordered_keys = list(color_map.keys())
    for key in ordered_keys:
        count = counter.get(key, 0)
        percent = round(100 * count / total, 1)
        ws_data.append([key, count, percent])

    ws_data.cell(row=ws_data.max_row + 1, column=1, value="Report generated on")
    ws_data.cell(row=ws_data.max_row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    for col_idx in range(1, 4):
        max_len = max(len(str(cell.value)) for cell in ws_data[get_column_letter(col_idx)] if cell.value)
        ws_data.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    chart = PieChart()
    chart.title = "Status Distribution"
    chart.width = 20
    chart.height = 12
    chart.legend.position = 'r'

    cats = Reference(ws_data, min_col=1, min_row=2, max_row=9)
    vals = Reference(ws_data, min_col=2, min_row=2, max_row=9)
    chart.add_data(vals, titles_from_data=False)
    chart.set_categories(cats)

    series = chart.series[0]
    for idx, key in enumerate(ordered_keys):
        if key in color_map:
            dp = DataPoint(idx=idx, spPr=GraphicalProperties(solidFill=ColorChoice(srgbClr=color_map[key])))
            series.dPt.append(dp)

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.plot_area.layout = Layout(manualLayout=ManualLayout(x=0.00, y=0.20, w=0.50, h=0.60))
    ws.add_chart(chart, "F2")

    report_path = os.path.join(profiling_dir, f"report_{provider}.xlsx")
    wb.save(report_path)
    print("Report generated:", report_path)
    print(f" • Profiling JSON files are in '{profiling_dir}/'")
    print(f" • Optimized models are in '{models_dir}/'")


# --------------------------------------------------
# Excel Report Generation: Aggregated
# --------------------------------------------------

def generate_report_aggregated(results, provider, profiling_dir, models_dir):
    """
    Generates an Excel file report_{provider}_aggregated.xlsx from the `results` list.
    Aggregated categories:
      - SUCCESS
      - SUCCESS WITH FALLBACK
      - UNKNOWN (no Node event)
      - NOT TESTED
      - SKIPPED
      - FAIL
    """

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(["op", "provider", "used_provider", "status"])

    agg_color_map = {
        "SUCCESS": "00AA44",
        "SUCCESS WITH FALLBACK": "FFAA00",
        "UNKNOWN (no Node event)": "DEDEDE",
        "NOT TESTED": "4D7CFE",
        "SKIPPED": "CCCCCC",
        "FAIL": "FF0000",
    }

    def get_fill_for_agg_status(status):
        normalized = "SUCCESS" if status.startswith("SUCCESS (with complexification)") else status
        if normalized.startswith("SUCCESS") and "FALLBACK" not in normalized:
            return PatternFill("solid", fgColor=agg_color_map["SUCCESS"])
        elif "WITH FALLBACK" in normalized:
            return PatternFill("solid", fgColor=agg_color_map["SUCCESS WITH FALLBACK"])
        elif normalized.startswith("UNKNOWN"):
            return PatternFill("solid", fgColor=agg_color_map["UNKNOWN (no Node event)"])
        elif normalized.startswith("NOT TESTED"):
            return PatternFill("solid", fgColor=agg_color_map["NOT TESTED"])
        elif normalized.startswith("SKIPPED"):
            return PatternFill("solid", fgColor=agg_color_map["SKIPPED"])
        elif normalized.startswith("FAIL"):
            return PatternFill("solid", fgColor=agg_color_map["FAIL"])
        return PatternFill("solid", fgColor=agg_color_map["FAIL"])

    for row in results:
        op_name, provider_used, used_provider, status = row
        display_status = "SUCCESS" if status.startswith("SUCCESS (with complexification)") else status
        ws.append([op_name, provider_used, used_provider or "", display_status])
        ws.cell(row=ws.max_row, column=4).fill = get_fill_for_agg_status(display_status)

    for col_idx in range(1, 5):
        max_len = max(len(str(cell.value)) for cell in ws[get_column_letter(col_idx)] if cell.value)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)

    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = border

    counter_agg = Counter()
    for _, _, _, status in results:
        normalized = "SUCCESS" if status.startswith("SUCCESS (with complexification)") else status
        if normalized.startswith("SUCCESS") and "FALLBACK" not in normalized:
            counter_agg["SUCCESS"] += 1
        elif "WITH FALLBACK" in normalized:
            counter_agg["SUCCESS WITH FALLBACK"] += 1
        elif normalized.startswith("UNKNOWN"):
            counter_agg["UNKNOWN (no Node event)"] += 1
        elif normalized.startswith("NOT TESTED"):
            counter_agg["NOT TESTED"] += 1
        elif normalized.startswith("SKIPPED"):
            counter_agg["SKIPPED"] += 1
        elif normalized.startswith("FAIL"):
            counter_agg["FAIL"] += 1
        else:
            counter_agg["FAIL"] += 1

    ws_data = wb.create_sheet("Data_PieChart")
    ws_data.append(["Category", "Count", "Percentage"])

    ordered_keys = list(agg_color_map.keys())
    total_agg = sum(counter_agg.values()) or 1
    for key in ordered_keys:
        count = counter_agg.get(key, 0)
        percent = round(100 * count / total_agg, 1)
        ws_data.append([key, count, percent])

    ts_row = ws_data.max_row + 1
    ws_data.cell(row=ts_row, column=1, value="Report generated on")
    ws_data.cell(row=ts_row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    for col_idx in range(1, 4):
        max_len = max(len(str(cell.value)) for cell in ws_data[get_column_letter(col_idx)] if cell.value)
        ws_data.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    chart = PieChart()
    chart.title = "Status Distribution (aggregated)"
    chart.width = 20
    chart.height = 12
    chart.legend.position = 'r'

    cats = Reference(ws_data, min_col=1, min_row=2, max_row=7)
    vals = Reference(ws_data, min_col=2, min_row=2, max_row=7)
    chart.add_data(vals, titles_from_data=False)
    chart.set_categories(cats)

    series = chart.series[0]
    for idx, key in enumerate(ordered_keys):
        if key in agg_color_map:
            dp = DataPoint(idx=idx, spPr=GraphicalProperties(solidFill=ColorChoice(srgbClr=agg_color_map[key])))
            series.dPt.append(dp)

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.plot_area.layout = Layout(manualLayout=ManualLayout(x=0.00, y=0.20, w=0.50, h=0.60))
    ws.add_chart(chart, "F2")

    report_path = os.path.join(profiling_dir, f"report_{provider}_aggregated.xlsx")
    wb.save(report_path)

    print("Aggregated report generated:", report_path)
    print(f" • Profiling JSON files are in '{profiling_dir}/'")
    print(f" • Optimized models are in '{models_dir}/'")



# --------------------------------------------------
# Utility functions to retrieve cuDNN and TensorRT versions
# --------------------------------------------------

def parse_cudnn_header(header_path):
    """
    Reads the cudnn_version.h or cudnn.h file and extracts CUDNN_MAJOR, CUDNN_MINOR, CUDNN_PATCHLEVEL.
    Returns "X.Y.Z" if found, otherwise None.
    """
    try:
        with open(header_path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()
    except Exception:
        return None

    major = re.search(r"#define\s+CUDNN_MAJOR\s+(\d+)", content)
    minor = re.search(r"#define\s+CUDNN_MINOR\s+(\d+)", content)
    patch = re.search(r"#define\s+CUDNN_PATCHLEVEL\s+(\d+)", content)
    if major and minor and patch:
        return f"{major.group(1)}.{minor.group(1)}.{patch.group(1)}"
    return None


def get_cudnn_version():
    """
    Searches multiple possible locations to find cudnn_version.h or cudnn.h under the CUDA toolkit
    and returns the version formatted as "X.Y.Z" if found.
    """
    cuda_candidates = []

    # 1) If CUDA_PATH is set, add it
    cuda_env = os.environ.get("CUDA_PATH")
    if cuda_env:
        cuda_candidates.append(cuda_env)
        # If CUDA_PATH ends with "vX.Y", also check the parent directory
        if os.path.basename(cuda_env).lower().startswith("v"):
            cuda_candidates.append(os.path.dirname(cuda_env))

    # 2) Standard installation path under Program Files on Windows
    prog_files = os.environ.get("ProgramFiles", r"C:\Program Files")
    default_root = os.path.join(prog_files, "NVIDIA GPU Computing Toolkit", "CUDA")
    cuda_candidates.append(default_root)

    # 3) Iterate through each candidate directory to detect cudnn_version.h or cudnn.h
    for base in cuda_candidates:
        if not os.path.isdir(base):
            continue

        # 3.a) Check if base/include contains cudnn_version.h or cudnn.h
        include_dir = os.path.join(base, "include")
        if os.path.isdir(include_dir):
            for filename in ("cudnn_version.h", "cudnn.h"):
                path_h = os.path.join(include_dir, filename)
                if os.path.isfile(path_h):
                    version = parse_cudnn_header(path_h)
                    if version:
                        return version

        # 3.b) Otherwise, search each subdirectory (e.g., v11.8, v12.5, etc.)
        for sub in os.listdir(base):
            sub_include = os.path.join(base, sub, "include")
            if not os.path.isdir(sub_include):
                continue
            for filename in ("cudnn_version.h", "cudnn.h"):
                path_h = os.path.join(sub_include, filename)
                if os.path.isfile(path_h):
                    version = parse_cudnn_header(path_h)
                    if version:
                        return version

    return "Unknown"


def get_tensorrt_version():
    """
    Attempts to call 'trtexec' to retrieve the TensorRT version.
    - If 'trtexec' prints a string like 'TensorRT.trtexec [TensorRT v100900]',
      extract '100900' and convert it to '10.9.0'.
    - If 'trtexec --version' already returns 'TensorRT v10.9.0', return '10.9.0' directly.
    - On error or if nothing is found, return 'Unknown'.
    """
    # Allow either a dot or a space between "TensorRT" and "trtexec"
    version_pattern = re.compile(r"TensorRT[.\s]+trtexec\s+\[TensorRT\s+v([\d\.]+)\]", re.IGNORECASE)

    def normalize(raw: str) -> str:
        # If the string already contains dots, return it as-is
        if "." in raw:
            return raw
        # If it's a 6-digit number like "100900", split into major, minor, patch
        if raw.isdigit() and len(raw) == 6:
            major = int(raw[0:2])
            minor = int(raw[2:4])
            patch = int(raw[4:6])
            return f"{major}.{minor}.{patch}"
        return raw

    try:
        # 1) Call "trtexec" without check=True to allow non-zero return codes
        proc = subprocess.run(
            ["trtexec"],
            capture_output=True,
            text=True
        )
        output = proc.stdout + proc.stderr

        match = version_pattern.search(output)
        if match:
            return normalize(match.group(1))

        # 2) If not found, try with "--version"
        proc2 = subprocess.run(
            ["trtexec", "--version"],
            capture_output=True,
            text=True
        )
        output2 = proc2.stdout + proc2.stderr
        match2 = version_pattern.search(output2)
        if match2:
            return normalize(match2.group(1))

    except FileNotFoundError:
        return "Unknown"
    except Exception:
        return "Unknown"

    return "Unknown"


# --------------------------------------------------
# Main function to generate the README in English
# --------------------------------------------------


def generate_readme_split(results, provider, output_dir, training_status_map=None, opset_version=None):

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    onnx_ver = onnx.__version__
    ort_ver = ort.__version__

    try:
        cpu_name = cpuinfo.get_cpu_info().get("brand_raw", platform.processor())
    except Exception:
        cpu_name = platform.processor() or "Unknown"

    gpu_list = []
    try:
        res = subprocess.run(["nvidia-smi", "--query-gpu=name", "--format=csv,noheader"],
                             capture_output=True, text=True, check=True)
        gpu_list = [g.strip() for g in res.stdout.strip().split("\n") if g.strip()]
    except Exception:
        pass

    cuda_version = "Unknown"
    try:
        res = subprocess.run(["nvcc", "--version"], capture_output=True, text=True, check=True)
        for line in res.stdout.splitlines():
            if "release" in line:
                cuda_version = line.split("release")[-1].split(",")[0].strip()
                break
    except Exception:
        pass

    cudnn_version = get_cudnn_version()
    trt_version = get_tensorrt_version()

    install_cmd = {
        "CUDAExecutionProvider": "pip install onnxruntime-gpu",
        "OpenVINOExecutionProvider": "pip install onnxruntime-openvino\npip install openvino==2025.1.0",
        "DmlExecutionProvider": "pip install onnxruntime-directml",
        "TensorrtExecutionProvider": "manual build with CUDA 12.5, cuDNN 9.2.1, TensorRT 10.9.0.34",
        "DnnlExecutionProvider": "manual build from source (oneDNN included, no pre-install needed)"
    }
    install_info = install_cmd.get(provider, "Installation method not specified")

    def summarize(subset, tag):
        counts = defaultdict(int)
        for _, _, _, status in subset:
            if status.startswith("SUCCESS (with complexification)"):
                counts["Executable directly (SUCCESS with complexification)"] += 1
            elif status.startswith("SUCCESS") and "FALLBACK" not in status:
                counts["SUCCESS"] += 1
            elif "WITH FALLBACK" in status:
                counts["FALLBACK"] += 1
            elif status.startswith("UNKNOWN"):
                counts["UNKNOWN"] += 1
            elif status.startswith("NOT TESTED"):
                counts["NOT TESTED"] += 1
            elif status.startswith("SKIPPED"):
                counts["SKIPPED"] += 1
            elif status.startswith("FAIL"):
                counts["FAIL"] += 1
            else:
                counts["FAIL"] += 1

        total = sum(counts.values()) or 1
        pie_path = os.path.join(output_dir, f"stats_{provider}_{tag}.png")

        fig, ax = plt.subplots(figsize=(6, 6))
        labels = [f"{k}: {counts[k]}" for k in counts]
        colors_map = {
            "SUCCESS": "#00AA44",
            "Executable directly (SUCCESS with complexification)": "#2299DD",
            "FALLBACK": "#FFAA00",
            "UNKNOWN": "#DEDEDE",
            "NOT TESTED": "#4D7CFE",
            "SKIPPED": "#CCCCCC",
            "FAIL": "#FF0000"
        }
        colors = [colors_map.get(k, "#AAAAAA") for k in counts]

        if total > 0:
            ax.pie(counts.values(), labels=labels, autopct='%1.1f%%', startangle=140, colors=colors)
            ax.axis('equal')
        else:
            ax.text(0.5, 0.5, "No nodes tested", ha='center', va='center', fontsize=12, color='gray')
            ax.axis('off')

        plt.savefig(pie_path, bbox_inches="tight")
        plt.close(fig)

        return counts, total, pie_path
    
    def summarize_training_subset(status_map, subset_ops, tag):
        """
        Filtre status_map sur subset_ops et génère un camembert (SUCCESS/FAIL/SKIPPED).
        NOT IMPLEMENTED est agrégé dans FAIL. Retourne (counts, total, png_path).
        """
        if not status_map or not subset_ops:
            return None, 0, None

        from collections import Counter
        import matplotlib.pyplot as plt

        def bucket(s):
            if s is None:
                return "SKIPPED"  # prudence si absence
            if s.startswith("OK") or s == "SUCCESS":
                return "SUCCESS"
            if s.startswith("NOT TESTED") or s.startswith("NOT_TESTED"):
                return "NOT_TESTED"
            if s.startswith("SKIPPED"):
                return "SKIPPED"
            # NOT IMPLEMENTED + autres => FAIL
            return "FAIL"

        # filtre sur le sous-ensemble
        filtered = {k: v for k, v in status_map.items() if k in subset_ops}
        if not filtered:
            return None, 0, None

        counter = Counter(bucket(v) for v in filtered.values())
        total = sum(counter.values()) or 1

        ordered = ["SUCCESS", "FAIL", "SKIPPED", "NOT_TESTED"]
        labels = [f"{k}: {counter.get(k, 0)}" for k in ordered if counter.get(k, 0) > 0]
        values = [counter.get(k, 0) for k in ordered if counter.get(k, 0) > 0]
        colors_map = {
            "SUCCESS": "#00AA44",
            "FAIL": "#FF0000",
            "SKIPPED": "#CCCCCC",
            "NOT_TESTED": "#4D7CFE",
        }
        colors = [colors_map[k] for k in ordered if counter.get(k, 0) > 0]

        pie_path = os.path.join(output_dir, f"training_stats_{provider}_{tag}.png")

        fig, ax = plt.subplots(figsize=(6, 6))
        if sum(values) > 0:
            ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=140, colors=colors)
            ax.axis('equal')
        else:
            ax.text(0.5, 0.5, "No training results", ha='center', va='center', fontsize=12, color='gray')
            ax.axis('off')
        plt.savefig(pie_path, bbox_inches="tight")
        plt.close(fig)

        return counter, total, pie_path


    def render_section(title, subset, tag):
        has_training = training_status_map is not None
        if has_training:
            lines = [f"## {title}", "", "| ONNX Node | Status | Training |", "|:---------:|:------:|:--------:|"]
        else:
            lines = [f"## {title}", "", "| ONNX Node | Status |", "|:---------:|:------:|"]
    
        def badge(label):
            color = {
                "SUCCESS": "00AA44",
                "FALLBACK": "FFAA00",
                "FAIL": "FF0000",
                "NOT TESTED": "7777CC",
                "SKIPPED": "999999",
                "UNKNOWN": "AAAAAA",
                "OK": "00AA44",
                "NOT IMPLEMENTED": "7777CC"
            }.get(label, "000000")
            return f"![{label}](https://img.shields.io/badge/{label.replace(' ', '%20')}-{color}?style=flat&logoColor=white)"
    
        for op_name, _, _, status in subset:
            # Statut exécution (déjà en place)
            if status.startswith("SUCCESS (with complexification)"):
                label = "SUCCESS"
            elif status.startswith("SUCCESS WITH FALLBACK"):
                label = "FALLBACK"
            elif status.startswith("SUCCESS"):
                label = "SUCCESS"
            elif status.startswith("FAIL"):
                label = "FAIL"
            elif status.startswith("NOT TESTED"):
                label = "NOT TESTED"
            elif status.startswith("SKIPPED"):
                label = "SKIPPED"
            else:
                label = "UNKNOWN"
    
            exec_badge = badge(label)
            link = f"https://onnx.ai/onnx/operators/onnx__{op_name}.html"
    
            if has_training:
                t_raw = training_status_map.get(op_name)
                if t_raw is None:
                    t_label = "SKIPPED"  # ex: pas de training pour cet EP → SKIPPED par défaut
                else:
                    if t_raw.startswith("OK") or t_raw == "SUCCESS":
                        t_label = "SUCCESS"
                    elif t_raw.startswith("NOT TESTED") or t_raw.startswith("NOT_TESTED"):
                        t_label = "NOT TESTED"
                    elif t_raw.startswith("SKIPPED"):
                        t_label = "SKIPPED"
                    elif t_raw.startswith("NOT IMPLEMENTED"):
                        t_label = "FAIL"
                    elif t_raw.startswith("FAIL"):
                        t_label = "FAIL"
                    else:
                        t_label = "UNKNOWN"

                t_badge = badge(t_label)
                lines.append(f"| [`{op_name}`]({link}) | {exec_badge} | {t_badge} |")
            else:
                lines.append(f"| [`{op_name}`]({link}) | {exec_badge} |")
    
        counts, total, pie_path = summarize(subset, tag)
        lines += [
            "",
            f"### Inference Summary",
            f"- **Total nodes tested:** {total}",
            f"- **Executable directly (SUCCESS):** {counts['SUCCESS']} ({round((counts['SUCCESS']/total)*100, 1)}%)",
            f"- **Executable directly (SUCCESS with complexification):** {counts['Executable directly (SUCCESS with complexification)']} ({round((counts['Executable directly (SUCCESS with complexification)']/total)*100, 1)}%)",
            f"- **Executable via FALLBACK:** {counts['FALLBACK']} ({round((counts['FALLBACK']/total)*100, 1)}%)",
            f"- **UNKNOWN (no Node event):** {counts['UNKNOWN']} ({round((counts['UNKNOWN']/total)*100, 1)}%)",
            f"- **NOT TESTED:** {counts['NOT TESTED']} ({round((counts['NOT TESTED']/total)*100, 1)}%)",
            f"- **SKIPPED:** {counts['SKIPPED']} ({round((counts['SKIPPED']/total)*100, 1)}%)",
            f"- **FAIL:** {counts['FAIL']} ({round((counts['FAIL']/total)*100, 1)}%)",
            "",
            f"![Pie Chart](./{os.path.basename(pie_path)})",
            ""
        ]
        return lines


    basic_results = [r for r in results if not r[0].startswith("com.microsoft.")]
    ms_results = [r for r in results if r[0].startswith("com.microsoft.")]

    os.makedirs(output_dir, exist_ok=True)
    lines = []
    lines.append(f"# ONNXRuntime Test Results — Provider: `{provider}`\n")
    lines.append(f"**Test Date:** {now}\n")
    # --- Test Methodology (split: Inference / Training) ---
    lines.append("## Test Methodology\n")
    
    # Inference
    if provider == "OpenVINOExecutionProvider":
        lines.append("### Inference")
        lines.append(
            "Each ONNX operator is tested using a minimal ONNX model containing only that specific node.\n"
            "In cases where OpenVINO falls back to CPU for simple models, we re-test the node with a slightly complexified "
            "model. This model adds a chain of `Mul` or `And` operations (based on input type) that preserve the behavior "
            "but help OpenVINO recognize and execute the subgraph. This allows better detection of real OpenVINO support."
        )
        lines.append("")
    else:
        lines.append("### Inference")
        lines.append(
            "Each ONNX operator is tested individually using a minimal ONNX model containing only that specific node. "
            "This ensures a focused and isolated evaluation of operator support for the selected Execution Provider."
        )
        lines.append("")
    
    # Training
    lines.append("### Training")
    lines.append(
        "To validate training (backward) support with ONNX Runtime Training, we **inject a `Mul` node** just before the "
        "tested operator: the **first input** of the tested node is multiplied by a **trainable scalar** `__train_C` "
        "(initialized to **1.0** so the forward values remain unchanged). We focus on the first input because it generally "
        "carries the data flow; for symmetric binary ops (e.g., `Add`), if training works on the first path it usually "
        "works on the others as well.\n\n"
        "We then generate ONNX Runtime **training artifacts** (AdamW), run an inference once to **patch output shapes** if "
        "needed, feed a **target equal to the model’s own output** (MSE loss on the first output), and perform **one "
        "optimization step**. A node is marked **SUCCESS** when this step completes; **NOT_TESTED** for explicitly skipped ops "
        "(e.g., some recurrent ops like GRU/LSTM) or unsupported input types for this method; otherwise it is **FAIL**."
    )
    lines.append("")


    lines.append("### Test Configuration\n")
    lines.append(f"- **ONNX Opset version:** {opset_version}")
    lines.append("- **ONNX IR version:** 10")
    lines.append("- **Data types:** Only one type is tested per node. This is usually `float32`, unless the node does not support it — in which case a compatible type is selected.\n")
    lines.append("> **Note:** Some ONNX nodes may not be available on the selected Execution Provider (EP) for opset version 22. "
                 "This can lead to fallback behavior even though these nodes were supported in earlier opset versions. "
                 "This occurs because ONNX Runtime teams may not have implemented or updated certain operators for the latest opset. "
                 "As a result, test outcomes can vary depending on both the ONNX opset version and the ONNX Runtime version used.\n")

    lines.append("## Environment and Installation Details\n")
    lines.append(f"- **ONNX version:** {onnx_ver}")
    lines.append(f"- **ONNXRuntime version:** {ort_ver}")
    lines.append(f"- **Target provider:** {provider}")
    lines.append(f"- **Installation command:**\n```bash\n{install_info}\n```")
    lines.append("### Hardware and Software Versions\n")
    lines.append(f"- **CPU:** {cpu_name}")
    if provider in ["CUDAExecutionProvider", "TensorrtExecutionProvider"] and gpu_list:
        lines.append(f"- **GPU(s):** {', '.join(gpu_list)}")
        lines.append(f"- **CUDA version:** {cuda_version}")
        lines.append(f"- **cuDNN version:** {cudnn_version}")
        if provider == "TensorrtExecutionProvider":
            lines.append(f"- **TensorRT version:** {trt_version}")
    lines.append("")

    lines += render_section("Basic ONNX Nodes", basic_results, "basic")
    
    # --- Training summary pour Basic ONNX Nodes ---
    if training_status_map:
        basic_ops_names = [op_name for (op_name, _, _, _) in basic_results]
        tr_counts, tr_total, tr_pie = summarize_training_subset(training_status_map, basic_ops_names, "basic")
        if tr_counts:
            s_ok    = tr_counts.get("SUCCESS", 0)
            s_fail  = tr_counts.get("FAIL", 0)
            s_skip  = tr_counts.get("SKIPPED", 0)
            s_nt    = tr_counts.get("NOT_TESTED", 0)
            
            lines += [
                "### Training Summary",
                f"- **Total nodes tested :** {tr_total}",
                f"- **SUCCESS:** {s_ok}",
                f"- **FAIL:** {s_fail}",
                f"- **SKIPPED (inférence FAIL|FALLBACK / no attempt):** {s_skip}",
                f"- **NOT TESTED (GRU/LSTM because it crash python kernel):** {s_nt}",
                f"![]({tr_pie})" if tr_pie else "",
                ""
            ]

    lines += render_section("Microsoft Custom Nodes", ms_results, "ms")
    
    # --- Training summary pour Microsoft Custom Nodes ---
    if training_status_map:
        ms_ops_names = [op_name for (op_name, _, _, _) in ms_results]
        tr_counts, tr_total, tr_pie = summarize_training_subset(training_status_map, ms_ops_names, "ms")
        if tr_counts:
            s_ok    = tr_counts.get("SUCCESS", 0)
            s_fail  = tr_counts.get("FAIL", 0)
            s_skip  = tr_counts.get("SKIPPED", 0)
            s_nt    = tr_counts.get("NOT_TESTED", 0)
            
            lines += [
                "### Training Summary",
                f"- **Total nodes tested :** {tr_total}",
                f"- **SUCCESS:** {s_ok}",
                f"- **FAIL:** {s_fail}",
                f"- **SKIPPED (inférence FAIL|FALLBACK / no attempt):** {s_skip}",
                f"- **NOT TESTED (GRU/LSTM because it crash python kernel):** {s_nt}",
                f"![]({tr_pie})" if tr_pie else "",
                ""
            ]

    skipped = [op for op, _, _, st in results if "SKIPPED" in st]
    not_tested = [op for op, _, _, st in results if st.startswith("NOT TESTED")]

    if not_tested:
        lines.append("## Nodes not tested\n")
        lines.append("These nodes couldn't be tested due to lack of valid minimal ONNX model.\n")
        lines.append(", ".join(f"`{n}`" for n in sorted(not_tested)))
        lines.append("")
    if skipped:
        lines.append("### Nodes skipped due to runtime errors\n")
        lines.append(", ".join(f"`{n}`" for n in sorted(skipped)))
        lines.append("")

    lines.append("## README Generation\n")
    lines.append("This file was generated automatically by `report.py`.\n")
    lines.append("- Generated ONNX models: `models/<provider>/`")
    lines.append("- Profiling JSON files: `profiling/<provider>/`")
    lines.append("- Scripts: `main.py`, `report.py`, `utils.py`, `ops/*`")
    lines.append("_End of README_")

    readme_path = os.path.join(output_dir, "README.md")
    with open(readme_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"✅ Split README written to: {readme_path}")


def generate_full_readme_for_opset(opset: int):
    """
    Génère un README global **par opset** dans:
        <project_root>/opset_{opset}/README.md
    en agrégeant les README de chaque EP dans:
        <project_root>/opset_{opset}/<EP>/README.md
    """
    import urllib.parse

    source_dir   = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(source_dir)
    opset_dir    = os.path.join(project_root, f"opset_{opset}")
    output_path  = os.path.join(opset_dir, "README.md")

    if not os.path.isdir(opset_dir):
        print(f"[WARN] opset dir not found: {opset_dir}")
        return

    # --- Compteurs d'une section (ordonnés) ---
    # On inclut UNKNOWN dans le total (comme dans les README d’EP)
    regex = {
        "SUCCESS": re.compile(r"\*\*Executable directly \(SUCCESS\):\*\*\s*(\d+)", re.IGNORECASE),
        "SUCCESS (with complexification)": re.compile(r"\*\*Executable directly \(SUCCESS with complexification\):\*\*\s*(\d+)", re.IGNORECASE),
        "FALLBACK": re.compile(r"\*\*Executable via FALLBACK:\*\*\s*(\d+)", re.IGNORECASE),
        "UNKNOWN (no Node event)": re.compile(r"\*\*UNKNOWN \(no Node event\):\*\*\s*(\d+)", re.IGNORECASE),
        "NOT TESTED": re.compile(r"\*\*NOT TESTED:\*\*\s*(\d+)", re.IGNORECASE),
        "SKIPPED": re.compile(r"\*\*SKIPPED:\*\*\s*(\d+)", re.IGNORECASE),
        "FAIL": re.compile(r"\*\*FAIL:\*\*\s*(\d+)", re.IGNORECASE),
    }

    def parse_section(readme: str, title: str) -> list[int] | None:
        """
        Extrait les compteurs d'inférence UNIQUEMENT depuis le bloc '### Statistics'
        à l'intérieur de la section '## {title}'.
        On évite ainsi de compter les lignes du bloc '### Training Summary (...)'.
        """
        # 1) isole la section H2 exacte
        h2_pat = re.compile(rf"^##\s+{re.escape(title)}\s*$", re.IGNORECASE | re.MULTILINE)
        m = h2_pat.search(readme)
        if not m:
            return None
        sec_start = m.end()
        nxt_h2 = re.search(r"^##\s+", readme[sec_start:], re.MULTILINE)
        sec_end = sec_start + (nxt_h2.start() if nxt_h2 else 0) if nxt_h2 else len(readme)
        section = readme[sec_start:sec_end]
    
        # 2) à l'intérieur, isole le sous-bloc '### Statistics'
        h3_stats = re.compile(r"^###\s+(Inference Summary|Statistics)\s*$", re.IGNORECASE | re.MULTILINE)
        ms = h3_stats.search(section)
        if not ms:
            return None
        st_start = ms.end()
        nxt_h3_or_h2 = re.search(r"^(###|##)\s+", section[st_start:], re.MULTILINE)
        st_end = st_start + (nxt_h3_or_h2.start() if nxt_h3_or_h2 else 0) if nxt_h3_or_h2 else len(section)
        stats_block = section[st_start:st_end]
    
        # 3) compte les catégories dans ce bloc uniquement
        results = {k: 0 for k in regex}
        for k, pat in regex.items():
            for val in pat.findall(stats_block):
                results[k] += int(val)
    
        # Retourne 7 valeurs (SUCCESS, SUCCESS with complexification, FALLBACK, UNKNOWN, NOT TESTED, SKIPPED, FAIL)
        return [results[k] for k in regex]


    # --- Training : on ne l’affiche que pour CPU / CUDA, et séparément Basic/MS ---
    tr_re_total   = re.compile(r"\*\*Total nodes tested.*?:\*\*\s*(\d+)", re.IGNORECASE)
    tr_re_success = re.compile(r"\*\*SUCCESS:\*\*\s*(\d+)", re.IGNORECASE)

    def slice_training_block(readme: str, header_text: str) -> str:
        # 1) Essai strict: "### Training Summary (Basic|Microsoft)"
        pat = re.compile(rf"^###\s+Training Summary\s*\({re.escape(header_text)}\)\s*$",
                         re.IGNORECASE | re.MULTILINE)
        m = pat.search(readme)
        if m:
            start = m.end()
            nxt = re.search(r"^###\s+", readme[start:], re.MULTILINE)
            end = start + (nxt.start() if nxt else 0) if nxt else len(readme)
            return readme[m.start():end]
    
        # 2) Fallback: chercher dans la section H2 correspondante un "### Training Summary" générique
        h2_pat = re.compile(rf"^##\s+{re.escape(header_text)}\s*$", re.IGNORECASE | re.MULTILINE)
        mh2 = h2_pat.search(readme or "")
        if not mh2:
            return ""
        sec_start = mh2.end()
        nxt_h2 = re.compile(r"^##\s+", re.MULTILINE).search(readme, sec_start)
        sec_end = nxt_h2.start() if nxt_h2 else len(readme)
        section = readme[sec_start:sec_end]
    
        h3_pat = re.compile(r"^###\s+Training Summary\s*$", re.IGNORECASE | re.MULTILINE)
        mh3 = h3_pat.search(section)
        if not mh3:
            return ""
        st = mh3.end()
        nxt = re.compile(r"^(###|##)\s+", re.MULTILINE).search(section, st)
        ed = nxt.start() if nxt else len(section)
        return section[mh3.start():ed]


    def parse_training_counts(readme: str, kind: str) -> tuple[int, int]:
        header = "Basic ONNX Nodes" if kind == "basic" else "Microsoft Custom Nodes"
        block = slice_training_block(readme or "", header)
        if not block:
            return (0, 0)
        mt, ms = tr_re_total.search(block), tr_re_success.search(block)
        tot  = int(mt.group(1)) if mt else 0
        succ = int(ms.group(1)) if ms else 0
        if tot < succ:
            tot = succ
        return succ, tot

    # --- EP folders ---
    ep_folders = sorted(
        d for d in os.listdir(opset_dir)
        if os.path.isdir(os.path.join(opset_dir, d))
        and os.path.isfile(os.path.join(opset_dir, d, "README.md"))
    )

    rows_basic, rows_ms = [], []

    def fmt_row(arr7: list[int]) -> list[str]:
        # indices: 0=SUCCESS,1=SUCCESS(w/complex),2=FALLBACK,3=UNKNOWN,4=NOT TESTED,5=SKIPPED,6=FAIL
        if not arr7 or len(arr7) != 7:
            arr7 = (arr7 or []) + [0] * (7 - len(arr7 or []))
            arr7 = arr7[:7]
        total = sum(arr7) or 1
        success_total = arr7[0] + arr7[1]
        fallback = arr7[2]
        fail = arr7[6]
        not_tested = arr7[4]
        skipped = arr7[5]
        supported = success_total + fallback
        return [
            f"{success_total} ({round(100*success_total/total)}%)",
            f"{fallback} ({round(100*fallback/total)}%)",
            f"{supported} ({round(100*supported/total)}%)",
            f"{fail} ({round(100*fail/total)}%)",
            f"{not_tested} ({round(100*not_tested/total)}%)",
            f"{skipped} ({round(100*skipped/total)}%)",
        ]

    for ep in ep_folders:
        path = os.path.join(opset_dir, ep, "README.md")
        with open(path, encoding="utf-8") as f:
            content = f.read()

        basic = parse_section(content, "Basic ONNX Nodes") or [0] * 7
        ms    = parse_section(content, "Microsoft Custom Nodes") or [0] * 7

        # Training uniquement pour CPU/CUDA
        ep_l = ep.strip().lower()
        training_enabled = ("cpu" in ep_l) or ("cuda" in ep_l)
        tr_b_succ = tr_b_tot = tr_m_succ = tr_m_tot = 0
        if training_enabled:
            tr_b_succ, tr_b_tot = parse_training_counts(content, "basic")
            tr_m_succ, tr_m_tot = parse_training_counts(content, "ms")

        rows_basic.append(
            [ep] + fmt_row(basic)
            + [f"{tr_b_succ} ({round(100*tr_b_succ/tr_b_tot)}%)" if tr_b_tot else "0 (0%)"]
        )
        rows_ms.append(
            [ep] + fmt_row(ms)
            + [f"{tr_m_succ} ({round(100*tr_m_succ/tr_m_tot)}%)" if tr_m_tot else "0 (0%)"]
        )

    # ---------- Rendu ----------
    cpu_name = platform.processor() or "Unknown"
    try:
        import cpuinfo as _cpuinfo
        cpu_name = _cpuinfo.get_cpu_info().get("brand_raw", cpu_name)
    except Exception:
        pass

    gpu_list = []
    try:
        res = subprocess.run(["nvidia-smi", "--query-gpu=name", "--format=csv,noheader"],
                             capture_output=True, text=True)
        gpu_list = [l.strip() for l in res.stdout.strip().splitlines() if l.strip()]
    except Exception:
        pass

    def write_table(out, title, rows):
        out.write(f"<h3>{title}</h3>\n")
        out.write("""<table border="1" cellpadding="6" cellspacing="0">
  <thead>
    <tr>
      <th>Execution Provider</th>
      <th>SUCCESS</th>
      <th>FALLBACK</th>
      <th>SUPPORTED</th>
      <th>FAIL</th>
      <th>NOT TESTED</th>
      <th>SKIPPED</th>
      <th>TRAINING</th>
    </tr>
  </thead>
  <tbody>
""")
        for row in rows:
            ep = row[0]
            ep_link = f'<a href="./{urllib.parse.quote(ep)}/" target="_blank">{ep}</a>'
            out.write(f"<tr><td>{ep_link}</td>"
                      f"<td>{row[1]}</td><td>{row[2]}</td><td>{row[3]}</td>"
                      f"<td>{row[4]}</td><td>{row[5]}</td><td>{row[6]}</td>"
                      f"<td><strong>{row[7]}</strong></td></tr>\n")
        out.write("</tbody></table>\n")

    # Ecriture du README d’opset
    with open(output_path, "w", encoding="utf-8") as out:
        out.write(f"""<div style="font-family:Arial, sans-serif; line-height:1.6; max-width:900px; margin:auto; padding:20px;">

<p align="center">
  <img src="https://github.com/microsoft/onnxruntime/raw/main/docs/images/ONNX_Runtime_logo_dark.png" alt="ONNX Runtime Logo" width="320"/>
</p>

<h1>ONNX Runtime — EP Coverage (Opset {opset})</h1>

<h2>🧪 What’s Tested</h2>
<ul>
  <li>Each ONNX operator is tested in isolation across all available EPs.</li>
  <li>Status per operator: <code>SUCCESS</code>, <code>FALLBACK</code>, <code>FAIL</code>, <code>NOT TESTED</code>, <code>SKIPPED</code>, <code>UNKNOWN</code>.</li>
  <li>Per-EP datasets include logs, optional optimized models, and a README with details.</li>
</ul>

<h2>📐 How’s Tested</h2>
<h3>Inference</h3>
<p>
  Minimal one-node ONNX model per op. A small “complexification” (e.g., extra <code>Mul</code>/<code>And</code>)
  can be added to trigger some compilers (OpenVINO/TensorRT) and reveal actual EP coverage.
</p>
<h3>Training</h3>
<p>
  When available (CPU/CUDA), a trainable scalar is injected before the tested node and a 1-step optimization (AdamW, MSE)
  validates basic backward. The training result appears only in the last column; it does not affect inference percentages.
</p>

<h2>📦 EPs with results in this opset</h2>
<ul>
""")
        for row in rows_basic:
            ep = row[0]
            out.write(f'<li><a href="./{urllib.parse.quote(ep)}/" target="_blank">{ep}</a></li>\n')
        out.write("</ul>\n")

        out.write(f"""
<h2>System / Versions</h2>
<ul>
  <li><strong>CPU:</strong> {cpu_name}</li>
  <li><strong>GPU:</strong> {', '.join(gpu_list) if gpu_list else 'No NVIDIA GPU detected'}</li>
  <li><strong>ONNX:</strong> {onnx.__version__} | <strong>ONNX Runtime:</strong> {ort.__version__}</li>
  <li><strong>ONNX Opset:</strong> {opset} | <strong>ONNX IR:</strong> 10</li>
</ul>
""")

        write_table(out, "ONNX Core Operators", rows_basic)
        write_table(out, "Microsoft Custom Operators", rows_ms)

        out.write("""
<h2>🧭 Related Tools</h2>
<p>
  For a complementary and more aggregated perspective on backend compliance, we encourage you to also visit the official 
  <a href="https://onnx.ai/backend-scoreboard/" target="_blank"><strong>ONNX Backend Scoreboard</strong></a>.
</p>
<p>
  While the Scoreboard provides a high-level view of backend support based on ONNX's internal test suite, our initiative focuses 
  on operator-level validation and runtime behavior analysis — especially fallback detection — across Execution Providers. 
  Together, both efforts help build a clearer, more actionable picture of ONNX Runtime capabilities.
</p>

<h2>🤝 Maintainer</h2>
<p>
  This project is maintained by <strong><a href="https://graiphic.io/" target="_blank">Graiphic</a></strong> 
  as part of the <a href="https://graiphic.io/download/" target="_blank"><strong>SOTA</strong></a> initiative.
</p>
<p>
  We welcome collaboration, community feedback, and open contribution to make ONNX Runtime stronger and more widely adopted.
</p>

<p style="margin-top:20px;">
  📬 <strong>Contact:</strong> <a href="mailto:contact@graiphic.io">contact@graiphic.io</a><br>
  🌐 <strong>Website:</strong> <a href="https://graiphic.io/" target="_blank">graiphic.io</a><br>
  🧠 <strong>Learn more about SOTA:</strong> <a href="https://graiphic.io/download/" target="_blank">graiphic.io/download</a>
</p>


</div>""")

    print(f"✅ Full README (opset {opset}) written to: {output_path}")



def generate_full_readme(opsets: list[int] | None = None):
    """
    Compat : si `opsets` est fourni -> génère un README dans chaque 'opset_<N>/'.
    Sinon, auto-découvre tous les dossiers 'opset_*'.
    """
    source_dir   = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(source_dir)

    if opsets is None:
        opsets = []
        for d in sorted(os.listdir(project_root)):
            if d.startswith("opset_"):
                try:
                    opsets.append(int(d.split("_", 1)[1]))
                except Exception:
                    pass

    for opset in sorted(set(opsets)):
        generate_full_readme_for_opset(opset)


def generate_root_readme():
    """
    Génère un README **global** à la racine du repo (Execution Providers Tester/README.md)
    en conservant toutes les sections de l'ancien README global SAUF :
      - 'Currently Supported Execution Providers'
      - 'Summary of ONNX Execution Provider Results'
    Un court paragraphe renvoie vers 'opset_*/README.md' pour les détails.
    """
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    output_path  = os.path.join(project_root, "README.md")

    with open(output_path, "w", encoding="utf-8") as out:
        out.write(f"""<div style="font-family:Arial, sans-serif; line-height:1.6; max-width:900px; margin:auto; padding:20px;">

<p align="center">
  <img src="https://github.com/microsoft/onnxruntime/raw/main/docs/images/ONNX_Runtime_logo_dark.png" alt="ONNX Runtime Logo" width="320"/>
</p>

<h1>Welcome to the ONNX Runtime – Execution Provider Coverage Tester</h1>

<p>
  This open source initiative, led by <strong><a href="https://graiphic.io/" target="_blank">Graiphic</a></strong>, provides 
  a detailed, real-world coverage map of ONNX operator support for each <strong>Execution Provider (EP)</strong> in 
  <strong><a href="https://github.com/microsoft/onnxruntime" target="_blank">ONNX Runtime</a></strong>.
</p>

<p>
  It is part of our broader effort to democratize AI deployment through 
  <a href="https://graiphic.io/download/" target="_blank"><strong>SOTA</strong></a> — 
  an ONNX-native orchestration framework designed for engineers, researchers, and industrial use cases.
</p>

<h2>🎯 Project Objectives</h2>
<ul>
  <li>Systematically test and report ONNX operator coverage per Execution Provider.</li>
  <li>Deliver up-to-date insights to guide industrial and academic ONNX Runtime adoption.</li>
  <li>Help developers, maintainers, and hardware vendors prioritize missing or broken operator support.</li>
</ul>

<h2>🧪 What’s Tested</h2>
<ul>
  <li>Each ONNX operator is tested in isolation using a minimal single-node model.</li>
  <li>Status per operator: <code>SUCCESS</code>, <code>FALLBACK</code>, <code>FAIL</code>, <code>NOT TESTED</code>, <code>SKIPPED</code>, <code>UNKNOWN</code>.</li>
  <li>Per-EP datasets include logs, optimized models (when applicable), and a README.</li>
</ul>

<h2>📐 How’s Tested</h2>
<h3>Inference</h3>
<p>
  Each operator is tested with a minimal ONNX graph. For EPs like OpenVINO/TensorRT, a <em>complexification</em> pass can add a small chain
  of <code>Mul</code>/<code>And</code> nodes (type-dependent) to make the backend compile more of the graph and reveal actual EP coverage.
</p>
<h3>Training</h3>
<p>
  When ONNX Runtime Training is available, a trainable scalar <code>__train_C</code> is injected via a <code>Mul</code> on the first input of the tested node (initialized to 1.0).
  We generate artifacts (AdamW) and run a single optimization step with an MSE loss on the first output. Operators that complete this step are marked <strong>SUCCESS</strong>;
  explicitly skipped or unsupported patterns are <strong>SKIPPED</strong>; others are <strong>FAIL</strong>.
</p>

<p><em>For detailed results and EP lists, please navigate to the per-opset dashboards:</em></p>
<ul>
""")
        # liste des dossiers opset_*/ existants
        for d in sorted(os.listdir(project_root)):
            if d.startswith("opset_") and os.path.isdir(os.path.join(project_root, d)):
                out.write(f'  <li><a href="./{d}/" target="_blank">{d}</a></li>\n')

        out.write("""</ul>

<h2>🧭 Related Tools</h2>
<p>
  For a complementary and more aggregated perspective on backend compliance, we encourage you to also visit the official 
  <a href="https://onnx.ai/backend-scoreboard/" target="_blank"><strong>ONNX Backend Scoreboard</strong></a>.
</p>
<p>
  While the Scoreboard provides a high-level view of backend support based on ONNX's internal test suite, our initiative focuses 
  on operator-level validation and runtime behavior analysis — especially fallback detection — across Execution Providers. 
  Together, both efforts help build a clearer, more actionable picture of ONNX Runtime capabilities.
</p>

<h2>🤝 Maintainer</h2>
<p>
  This project is maintained by <strong><a href="https://graiphic.io/" target="_blank">Graiphic</a></strong> 
  as part of the <a href="https://graiphic.io/download/" target="_blank"><strong>SOTA</strong></a> initiative.
</p>
<p>
  We welcome collaboration, community feedback, and open contribution to make ONNX Runtime stronger and more widely adopted.
</p>

<p style="margin-top:20px;">
  📬 <strong>Contact:</strong> <a href="mailto:contact@graiphic.io">contact@graiphic.io</a><br>
  🌐 <strong>Website:</strong> <a href="https://graiphic.io/" target="_blank">graiphic.io</a><br>
  🧠 <strong>Learn more about SOTA:</strong> <a href="https://graiphic.io/download/" target="_blank">graiphic.io/download</a>
</p>

</div>""")

    print(f"✅ Global root README written to: {output_path}")
