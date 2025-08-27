# utils.py — refactor clean & conditional training
import sys
from pathlib import Path
import os
import json
import pkgutil
import importlib
from collections import defaultdict


import numpy as np
import onnx
import onnx.helper
import onnxruntime as ort

from report import (
    generate_report,
    generate_report_aggregated,
    generate_readme_split,
)

# ---------------------------------------------------------------------
# Constants & registries
# ---------------------------------------------------------------------

ONNX_RUNTIME_IR_VERSION = 10
ONNX_OPSET_VERSION = 22

SpecialModelBuilders = {}
SpecialInputGenerators = {}

def _reset_registries():
    SpecialModelBuilders.clear()
    SpecialInputGenerators.clear()

def _reload_ops_package():
    """
    Décharge puis recharge tous les modules du package 'ops'.
    L'import de chaque module ré-enregistrera SpecialModelBuilders/InputGenerators.
    """
    pkg_name = "ops"
    to_delete = [m for m in list(sys.modules.keys()) if m == pkg_name or m.startswith(pkg_name + ".")]
    for m in to_delete:
        del sys.modules[m]
    importlib.invalidate_caches()
    importlib.import_module(pkg_name)
    pkg = sys.modules[pkg_name]
    for _, name, _ in pkgutil.iter_modules(pkg.__path__):
        importlib.import_module(f"{pkg_name}.{name}")

def configure_opset(opset: int):
    """
    Met à jour ONNX_OPSET_VERSION puis recharge 'ops/*' pour que les sous-modules
    relisent la constante via `from utils import ONNX_OPSET_VERSION`.
    """
    global ONNX_OPSET_VERSION
    ONNX_OPSET_VERSION = int(opset)
    _reset_registries()
    _reload_ops_package()


# ---------------------------------------------------------------------
# Training availability
# ---------------------------------------------------------------------

def has_ort_training() -> bool:
    """
    Returns True iff onnxruntime-training is available in this environment.
    """
    try:
        import onnxruntime.training as _ortt  # noqa: F401
        import onnxruntime.training.artifacts as _art  # noqa: F401
        return True
    except Exception:
        return False


# ---------------------------------------------------------------------
# Config inputs (skip / untested)
# ---------------------------------------------------------------------

def load_skip_table(path: str = "skip_nodes.json") -> dict:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

def load_not_tested_table(path: str = "untested_nodes.json") -> dict:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return {}


# ---------------------------------------------------------------------
# EP mapping (folder names)
# ---------------------------------------------------------------------

def map_execution_provider(provider_name: str) -> str:
    mapping = {
        "CPUExecutionProvider": "CPU",
        "CUDAExecutionProvider":       "NVIDIA - CUDA",
        "TensorrtExecutionProvider":   "NVIDIA - TensorRT",
        "OpenVINOExecutionProvider":   "Intel - OpenVINO™",
        "DnnlExecutionProvider":       "Intel - oneDNN",
        "DmlExecutionProvider":        "Windows - DirectML",
        "QNNExecutionProvider":        "Qualcomm - QNN",
        "NnapiExecutionProvider":      "Android - NNAPI",
        "CoreMLExecutionProvider":     "Apple - CoreML",
        "XnnpackExecutionProvider":    "XNNPACK",
        "ROCMExecutionProvider":       "AMD - ROCm",
        "MIGraphXExecutionProvider":   "AMD - MIGraphX",
        "VitisAIExecutionProvider":    "AMD - Vitis AI",
        "AzureExecutionProvider":      "Cloud - Azure",
        "ACLExecutionProvider":        "Arm - ACL",
        "ArmNNExecutionProvider":      "Arm - Arm NN",
        "TVMExecutionProvider":        "Apache - TVM",
        "RknpuExecutionProvider":      "Rockchip - RKNPU",
        "CANNExecutionProvider":       "Huawei - CANN",
    }
    return mapping.get(provider_name, "Unknown")


# ---------------------------------------------------------------------
# Default builders & generators
# ---------------------------------------------------------------------

def default_model_builder(op_type: str, cfg=None) -> onnx.ModelProto:
    """
    Creates a minimal ONNX model with a single node of type `op_type`.
    Input: tensor 'X' [1,1] float
    Output: 'Y' (shape inferred by ORT)
    """
    dtype = onnx.TensorProto.FLOAT
    inp = onnx.helper.make_tensor_value_info("X", dtype, [1, 1])
    out = onnx.helper.make_tensor_value_info("Y", dtype, None)
    node = onnx.helper.make_node(op_type, [inp.name], [out.name])
    graph = onnx.helper.make_graph([node], f"{op_type}_graph", [inp], [out])
    model = onnx.helper.make_model(
        graph, opset_imports=[onnx.helper.make_operatorsetid("", ONNX_OPSET_VERSION)]
    )
    model.ir_version = ONNX_RUNTIME_IR_VERSION
    return model


def default_input_generator(session: ort.InferenceSession) -> dict:
    """
    Generates random inputs for the given session.
    int64 -> randint; otherwise float32.
    Symbolic dims are replaced with 1.
    """
    feed = {}
    for inp in session.get_inputs():
        shape = [d if isinstance(d, int) else 1 for d in inp.shape]
        if inp.type == "tensor(int64)":
            feed[inp.name] = np.random.randint(0, 10, size=shape, dtype=np.int64)
        else:
            feed[inp.name] = np.random.rand(*shape).astype(np.float32)
    return feed


class OpTest:
    """
    Encapsulates building, saving, and running a single-operator ONNX model.
    """

    def __init__(self, op_type: str):
        self.op_type = op_type

    def generate_model(self) -> onnx.ModelProto:
        builder = SpecialModelBuilders.get(self.op_type, default_model_builder)
        return builder(self.op_type)

    def generate_input(self, session: ort.InferenceSession) -> dict:
        gen = SpecialInputGenerators.get(self.op_type, default_input_generator)
        return gen(session)

    def save_model(self, model: onnx.ModelProto, directory: str = "models") -> str:
        os.makedirs(directory, exist_ok=True)
        path = os.path.join(directory, f"{self.op_type}.onnx")
        onnx.save(model, path)
        return path


# ---------------------------------------------------------------------
# Test list loader
# ---------------------------------------------------------------------

def load_ops(path: str):
    """
    Returns two sorted lists:
      - basic_ops: onnx core operators
      - microsoft_ops: custom nodes with prefix 'com.microsoft.'
    Lines empty or starting with '#' are ignored.
    """
    basic_ops, microsoft_ops = [], []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            name = line.strip()
            if not name or name.startswith("#"):
                continue
            (microsoft_ops if name.startswith("com.microsoft.") else basic_ops).append(name)
    return sorted(basic_ops), sorted(microsoft_ops)


# ---------------------------------------------------------------------
# OpenVINO / TensorRT complexification helpers
# ---------------------------------------------------------------------

def complexify_model_for_openvino(model: onnx.ModelProto) -> onnx.ModelProto:
    """
    Adds a chain of Mul/And nodes on the first graph input to encourage
    OpenVINO/TensorRT to compile the subgraph fully in its own EP.
    """
    from onnx import helper, TensorProto

    graph = model.graph
    input0 = graph.input[0]
    dtype = input0.type.tensor_type.elem_type
    input_name = input0.name

    # new scalar input
    noise_name = "noise"
    noise_vi = helper.make_tensor_value_info(noise_name, dtype, [])
    graph.input.append(noise_vi)

    numeric_types = {
        TensorProto.FLOAT, TensorProto.FLOAT16, TensorProto.DOUBLE,
        TensorProto.INT8, TensorProto.INT16, TensorProto.INT32, TensorProto.INT64,
        TensorProto.UINT8, TensorProto.UINT16, TensorProto.UINT32, TensorProto.UINT64,
    }
    op_type = "Mul" if dtype in numeric_types else "And"

    new_nodes = []
    prev_out = None
    for i in range(10):
        in0 = input_name if i == 0 else prev_out
        out_name = f"{op_type.lower()}_{i}"
        node = helper.make_node(op_type, inputs=[in0, noise_name], outputs=[out_name])
        new_nodes.append(node)
        prev_out = out_name

    # rewire first input usages
    for node in graph.node:
        for idx, name in enumerate(node.input):
            if name == input_name:
                node.input[idx] = prev_out

    for node in reversed(new_nodes):
        graph.node.insert(0, node)

    return model


def executed_on_provider_strict(op_name: str, events: list, provider: str) -> str:
    """
    Determines strict execution status for a given op and provider:
      - SUCCESS (all nodes on provider and op present)
      - SUCCESS WITH FALLBACK (some nodes on other EPs)
      - SUCCESS (with complexification) (no Node event for op but all nodes on provider)
    """
    nodes = [e for e in events if e.get("cat") == "Node"]

    fallback_eps = {
        e["args"]["provider"]
        for e in nodes
        if e["args"].get("provider") != provider
        and e["args"].get("op_name") not in ("MemcpyFromHost", "MemcpyToHost")
    }
    if fallback_eps:
        return "SUCCESS WITH FALLBACK"

    # op explicitly present on provider
    if any(e["args"].get("op_name") == op_name and e["args"]["provider"] == provider for e in nodes):
        return "SUCCESS"

    # otherwise consider as success via complexification
    return "SUCCESS (with complexification)"


# ---------------------------------------------------------------------
# Training graph injections (kept minimal & local imports)
# ---------------------------------------------------------------------

def add_training_scale(model: onnx.ModelProto) -> onnx.ModelProto:
    """
    Injects a trainable scalar (__train_C = 1.0) via a Mul on the first input
    of the first node. Only supports float-like first inputs.
    """
    from onnx import helper, numpy_helper, TensorProto

    g = model.graph
    if not g.node:
        raise ValueError("Empty model.")

    target = g.node[0]
    if not target.input:
        raise ValueError(f"Node {target.op_type} has no editable input.")

    original = target.input[0]
    inp_vi = next((i for i in g.input if i.name == original), None)
    if inp_vi is None:
        # we only handle raw graph inputs as first input for training injection
        raise ValueError(f"Cannot find graph.input for '{original}'.")

    elem_type = inp_vi.type.tensor_type.elem_type
    float_types = {
        onnx.TensorProto.FLOAT, onnx.TensorProto.FLOAT16,
        onnx.TensorProto.BFLOAT16, onnx.TensorProto.DOUBLE
    }
    if elem_type not in float_types:
        raise NotImplementedError(
            f"Training scaling unsupported for non-float first input (got {onnx.TensorProto.DataType.Name(elem_type)})."
        )

    scale_name = "__train_C"
    g.initializer.extend([numpy_helper.from_array(np.array(1.0, dtype=np.float32), scale_name)])

    mul_out = f"{original}_scaled"
    mul_node = helper.make_node("Mul", [original, scale_name], [mul_out])

    # rewire target first input
    target.input[0] = mul_out
    # insert Mul at the beginning
    g.node.insert(0, mul_node)

    # prune graph.input that are actually initializers
    init_names = {init.name for init in g.initializer}
    new_inputs = [vi for vi in g.input if vi.name not in init_names]
    del g.input[:]
    g.input.extend(new_inputs)

    # ensure base opset is present
    if "" not in {op.domain for op in model.opset_import}:
        model.opset_import.extend([helper.make_operatorsetid("", ONNX_OPSET_VERSION)])
    return model


# ---------------------------------------------------------------------
# Training Excel (SUCCESS/FAIL/SKIPPED only)
# ---------------------------------------------------------------------

def write_training_results_to_excel(results, filename="training_results.xlsx"):
    """
    results: list[(op, status, msg)] where status can be "OK", "FAIL ...", "SKIPPED", etc.
    Produces an Excel file with a table and a pie chart sheet (SUCCESS/FAIL/SKIPPED).
    """
    from collections import Counter
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.chart import Reference
    from openpyxl.chart.pie_chart import PieChart
    from openpyxl.chart.series import DataPoint, GraphicalProperties
    from openpyxl.chart.label import DataLabelList
    from openpyxl.drawing.colors import ColorChoice
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Training Results"
    ws.append(["Operator", "Status", "Message"])

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray  = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")   # SKIPPED
    blue  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")   # NOT_TESTED

    def bucket(s):
        if s.startswith("OK") or s == "SUCCESS":
            return "SUCCESS"
        if s.startswith("NOT_TESTED") or s.startswith("NOT TESTED"):
            return "NOT_TESTED"
        if s.startswith("SKIPPED"):
            return "SKIPPED"
        return "FAIL"


    counts = Counter()
    for op, status, msg in results:
        ws.append([op, status, msg])
        b = bucket(status)
        counts[b] += 1
        fill = green if b == "SUCCESS" else red if b == "FAIL" else blue if b == "NOT_TESTED" else gray
        for c in ws[ws.max_row]:
            c.fill = fill

    # autosize
    for col in range(1, 4):
        max_len = max(len(str(cell.value)) for cell in ws[get_column_letter(col)] if cell.value)
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 100)

    # Pie sheet
    ws_data = wb.create_sheet("Data_PieChart")
    ws_data.append(["Category", "Count", "Percentage"])
    ordered = ["SUCCESS", "FAIL", "SKIPPED", "NOT_TESTED"]  # <<< 4 catégories
    total = sum(counts.values()) or 1
    for k in ordered:
        c = counts.get(k, 0)
        p = round(100 * c / total, 1)
        ws_data.append([k, c, p])

    color_map = {"SUCCESS": "00AA44", "FAIL": "FF0000", "SKIPPED": "CCCCCC", "NOT_TESTED": "4D7CFE"}
    chart = PieChart()
    chart.title = "Training Status Distribution"
    chart.width = 20
    chart.height = 12
    chart.legend.position = "r"

    cats = Reference(ws_data, min_col=1, min_row=2, max_row=1 + len(ordered))
    vals = Reference(ws_data, min_col=2, min_row=2, max_row=1 + len(ordered))
    chart.add_data(vals, titles_from_data=False)
    chart.set_categories(cats)

    series = chart.series[0]
    for i, k in enumerate(ordered, start=2):
        c = counts.get(k, 0)
        if c == 0:
            continue
        dp = DataPoint(idx=i-2)
        dp.graphicalProperties = GraphicalProperties(solidFill=color_map[k])
        series.dPt.append(dp)

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    ws.add_chart(chart, "E2")

    wb.save(filename)
    print(f"📊 Training Excel saved: {filename}")


# ---------------------------------------------------------------------
# Training runner (isolated imports inside)
# ---------------------------------------------------------------------

def _run_training_for_provider(
    provider: str,
    ops_to_test: list,
    models_dir: str,
    base_output_dir: str,
    artifacts_root: str,
    inference_outcome_map: dict | None = None,  # <<< renommé / tri-état
):
    """
    Minimal 1-step training per op by injecting 'Scale' (Mul with __train_C).
    Only runs on CPU/CUDA if onnxruntime-training is present.
    Returns:
      results: list[(op, status, msg)]
      status_map: dict[op] -> "SUCCESS" | "FAIL" | "SKIPPED"
    """
    # provider filter first
    if provider not in ("CPUExecutionProvider", "CUDAExecutionProvider"):
        return [], {}

    # env availability
    if not has_ort_training():
        return [], {}

    # late imports (only if available)
    from onnx import TensorProto
    import onnxruntime as ort
    import onnxruntime.training.artifacts as artifacts
    from onnxruntime.training.api import Module, CheckpointState, Optimizer

    # Local definition to avoid import at module-import time
    try:
        import onnxruntime.training.onnxblock as onnxblock

        class FirstOutputMSELoss(onnxblock.Block):
            def __init__(self, target_name: str = "target"):
                super().__init__()
                self._mse = onnxblock.loss.MSELoss()
                self._target = target_name
            def build(self, first_output_name, *other_output_names):
                return self._mse(first_output_name, target_name=self._target)
    except Exception as e:
        print(f"[training] onnxblock not available: {e}")
        return [], {}

    os.makedirs(artifacts_root, exist_ok=True)
    device = "cuda" if provider == "CUDAExecutionProvider" else "cpu"

    results = []
    status_map = {}
    skip_ops = {"GRU", "LSTM"}  # explicit RNN skip

    def _force_output_shapes(model: onnx.ModelProto, outputs: list):
        """
        Patch graph.output shapes explicitly based on a sample run outputs.
        """
        def _set(vi, shp):
            vi.type.tensor_type.elem_type = vi.type.tensor_type.elem_type or TensorProto.FLOAT
            vi.type.tensor_type.ClearField("shape")
            for d in shp:
                dim = vi.type.tensor_type.shape.dim.add()
                if d is None:
                    dim.dim_param = "unk"
                else:
                    dim.dim_value = int(d)

        if len(model.graph.output) == len(outputs):
            for vi, arr in zip(model.graph.output, outputs):
                if not vi.type.tensor_type.HasField("shape"):
                    _set(vi, arr.shape)

    for op in ops_to_test:
        if op in skip_ops:
            results.append((op, "NOT_TESTED", "Not tested because it crash python kernel"))
            status_map[op] = "NOT_TESTED"
            continue
        # 2.2) Si l’inférence a échoué pour cet op, on n’essaie PAS le training
        if inference_outcome_map is not None:
            outcome = inference_outcome_map.get(op, "FAIL")
        if outcome == "FAIL":
            results.append((op, "SKIPPED", "Training not attempted: inference FAIL"))
            status_map[op] = "SKIPPED"
            continue
        if outcome == "SUCCESS_WITH_FALLBACK":
            results.append((op, "SKIPPED", "Training not attempted: inference FALLBACK"))
            status_map[op] = "SKIPPED"
            continue

        tester = OpTest(op)

        # 1) build base + inject scale
        try:
            base_model = tester.generate_model()
            model = add_training_scale(base_model)
            # infer sample output shapes on CPU
            tmp_sess = ort.InferenceSession(model.SerializeToString(), providers=["CPUExecutionProvider"])
            feed_shape = tester.generate_input(tmp_sess)
            outs = tmp_sess.run(None, feed_shape)
            _force_output_shapes(model, outs)
            onnx.save(model, os.path.join(models_dir, f"{op}_train.onnx"))
        except NotImplementedError as e:
            results.append((op, "FAIL not implemented", f"{e}"))
            status_map[op] = "FAIL"
            continue
        except Exception as e:
            results.append((op, "FAIL model gen", f"{e}"))
            status_map[op] = "FAIL"
            continue

        # 2) artifacts & single step
        try:
            art_dir = os.path.join(artifacts_root, op)
            os.makedirs(art_dir, exist_ok=True)

            loss_block = FirstOutputMSELoss(target_name="target")
            artifacts.generate_artifacts(
                model=model,
                requires_grad=["__train_C"],
                loss=loss_block,
                optimizer=artifacts.OptimType.AdamW,
                artifact_directory=art_dir,
            )

            module = Module(
                train_model_uri=os.path.join(art_dir, "training_model.onnx"),
                eval_model_uri=os.path.join(art_dir, "eval_model.onnx"),
                state=CheckpointState.load_checkpoint(os.path.join(art_dir, "checkpoint")),
                device=device,
            )

            # create feed
            inf_sess = ort.InferenceSession(model.SerializeToString(), providers=["CPUExecutionProvider"])
            inputs = tester.generate_input(inf_sess)
            predicted = inf_sess.run(None, inputs)[0].astype(np.float32)

            # detect expected non-initializer inputs for training_model
            train_model = onnx.load(os.path.join(art_dir, "training_model.onnx"))
            init_names = {init.name for init in train_model.graph.initializer}
            expected_inputs = [i.name for i in train_model.graph.input if i.name not in init_names]

            # find target name (prefer ...::target or 'target'; otherwise 1st unknown)
            target_name = None
            for cand in expected_inputs:
                if cand.endswith("::target") or cand == "target":
                    target_name = cand
                    break
            if target_name is None:
                for cand in expected_inputs:
                    if cand not in inputs:
                        target_name = cand
                        break

            feed_dict = dict(inputs)
            if target_name:
                feed_dict[target_name] = predicted

            ordered = [feed_dict[n] for n in expected_inputs if n in feed_dict]

            optimizer = Optimizer(os.path.join(art_dir, "optimizer_model.onnx"), module)
            module.train()
            module(*ordered)
            optimizer.step()
            module.lazy_reset_grad()

            results.append((op, "OK", "Training step completed"))
            status_map[op] = "SUCCESS"

        except Exception as e:
            msg = str(e)
            # All failures are reported as FAIL (you decided to merge NOT IMPLEMENTED into FAIL visually)
            results.append((op, "FAIL training", msg))
            status_map[op] = "FAIL"

    # write training Excel per provider folder
    xlsx_path = os.path.join(base_output_dir, f"training_{provider}.xlsx")
    write_training_results_to_excel(results, filename=xlsx_path)
    return results, status_map


# ---------------------------------------------------------------------
# Main test runner (inference + reports + optional training)
# ---------------------------------------------------------------------

def run_tests_and_generate_reports(
    provider: str = "DnnlExecutionProvider",
    test_file: str = "test.txt",
    profiling_dir: str = "profiling",
    models_dir: str = "models",
):
    """
    Runs single-node inference tests for the given provider, writes reports,
    then (if available and relevant) runs training checks and augments README.
    """
    is_openvino = "OpenVINO" in provider
    if is_openvino:
        try:
            import onnxruntime.tools.add_openvino_win_libs as ov_utils
            ov_utils.add_openvino_libs_to_path()
        except ImportError:
            print(f"[warn] OpenVINO helper not found for provider '{provider}'.")

    # clean profiling dir
    if os.path.exists(profiling_dir):
        for f in os.listdir(profiling_dir):
            p = os.path.join(profiling_dir, f)
            if os.path.isfile(p):
                os.remove(p)
    os.makedirs(profiling_dir, exist_ok=True)
    os.makedirs(models_dir, exist_ok=True)

    # load skip/not-tested once
    skip_table = load_skip_table()
    untested_table = load_not_tested_table()
    skip_list = set(skip_table.get("ALL", [])) | set(skip_table.get(provider, []))
    untested_list = set(untested_table.get("ALL", [])) | set(untested_table.get(provider, []))

    # auto-register ops/*
    import ops  # noqa: F401
    for _, name, _ in pkgutil.iter_modules(ops.__path__):
        importlib.import_module(f"ops.{name}")

    # gather ops to test
    basic_ops, microsoft_ops = load_ops(test_file)
    ops_to_test = basic_ops + microsoft_ops
    results = []

    # inference pass
    for op in ops_to_test:
        if op in skip_list:
            results.append((op, provider, None, "FAIL (skipped: known crash)"))
            continue
        if op in untested_list:
            results.append((op, provider, None, "NOT TESTED (model unavailable)"))
            continue

        tester = OpTest(op)

        # build + save
        try:
            model = tester.generate_model()
            tester.save_model(model, directory=models_dir)
        except Exception as e:
            results.append((op, provider, None, f"FAIL generate_model: {e}"))
            continue

        # session options
        opts = ort.SessionOptions()
        opts.enable_profiling = True
        opts.profile_file_prefix = os.path.join(profiling_dir, f"profile_{op}")
        opts.execution_mode = ort.ExecutionMode.ORT_SEQUENTIAL
        opts.log_severity_level = 4

        skip_optimized = provider in {"OpenVINOExecutionProvider", "TensorrtExecutionProvider", "DnnlExecutionProvider"}
        if not skip_optimized:
            opts.optimized_model_filepath = os.path.join(models_dir, f"{op}_optimized.onnx")

        # session creation (special cases)
        try:
            if is_openvino:
                sess = ort.InferenceSession(
                    model.SerializeToString(),
                    sess_options=opts,
                    providers=[provider],
                    provider_options=[{"device_type": "CPU"}],
                )
            elif provider == "TensorrtExecutionProvider":
                sess = ort.InferenceSession(
                    model.SerializeToString(),
                    sess_options=opts,
                    providers=[
                        "TensorrtExecutionProvider",
                        "CUDAExecutionProvider",
                        "CPUExecutionProvider",
                    ],
                )
            else:
                sess = ort.InferenceSession(
                    model.SerializeToString(),
                    sess_options=opts,
                    providers=[provider],
                )
        except Exception as e:
            results.append((op, provider, None, f"FAIL session creation: {e}"))
            continue

        # inputs
        try:
            feed = tester.generate_input(sess)
        except Exception as e:
            results.append((op, provider, None, f"FAIL generate_input: {e}"))
            continue

        # run
        try:
            iobinding_cb = feed.pop("__iobinding__", None)
            if iobinding_cb:
                iobinding_cb(sess)
            else:
                sess.run(None, feed)
            profile_path = sess.end_profiling()
        except Exception as e:
            results.append((op, provider, None, f"FAIL run: {e}"))
            continue

        # parse profile to detect EP usage
        try:
            with open(profile_path, "r") as f:
                root = json.load(f)
            events = root["traceEvents"] if isinstance(root, dict) and "traceEvents" in root else root
            op_events = [e for e in events if e.get("cat") == "Node" and e.get("args", {}).get("op_name") == op]

            if op_events:
                used = op_events[0]["args"]["provider"]
                status = "SUCCESS" if used == provider else "SUCCESS WITH FALLBACK"
            else:
                subops = [
                    e for e in events
                    if e.get("cat") == "Node"
                    and e.get("args", {}).get("op_name") not in ("MemcpyFromHost", "MemcpyToHost")
                ]
                if subops:
                    sub_providers = [e["args"]["provider"] for e in subops]
                    if all(p == provider for p in sub_providers):
                        used = provider
                        status = "SUCCESS (via decomposition)"
                    else:
                        used = ",".join(sorted(set(sub_providers)))
                        status = "SUCCESS WITH FALLBACK (via decomposition)"
                else:
                    used = None
                    status = "UNKNOWN (no Node event)"

            results.append((op, provider, used, status))
        except Exception as e:
            results.append((op, provider, None, f"FAIL parsing JSON: {e}"))
            continue

        # Optional complexification re-test for OpenVINO/TensorRT when fallback detected
        is_tensorrt = provider == "TensorrtExecutionProvider"
        if (is_openvino or is_tensorrt) and status.startswith("SUCCESS WITH FALLBACK"):
            try:
                feed2 = tester.generate_input(sess)
                if isinstance(feed2, tuple):
                    feed2 = feed2[0]
                if not isinstance(feed2, dict):
                    raise TypeError(f"generate_input for {op} must return a dict, got {type(feed2)}")

                complex_model = complexify_model_for_openvino(model)
                complex_path = os.path.join(models_dir, f"{op}_complex.onnx")
                onnx.save(complex_model, complex_path)
                opts2 = ort.SessionOptions()
                opts2.enable_profiling = True
                opts2.profile_file_prefix = os.path.join(profiling_dir, f"profile_{op}_complex")

                if is_openvino:
                    sess2 = ort.InferenceSession(
                        complex_model.SerializeToString(),
                        sess_options=opts2,
                        providers=[provider],
                        provider_options=[{"device_type": "CPU"}],
                    )
                else:
                    sess2 = ort.InferenceSession(
                        complex_model.SerializeToString(),
                        sess_options=opts2,
                        providers=["TensorrtExecutionProvider", "CUDAExecutionProvider", "CPUExecutionProvider"],
                    )

                # add 'noise' scalar input when required
                if "noise" in [i.name for i in sess2.get_inputs()]:
                    ref = next(iter(feed2.values()))
                    feed2["noise"] = np.array(1, dtype=ref.dtype)

                sess2.run(None, feed2)
                profile_path2 = sess2.end_profiling()
                with open(profile_path2, "r") as f:
                    root2 = json.load(f)
                events2 = root2["traceEvents"] if isinstance(root2, dict) else root2
                status2 = executed_on_provider_strict(op, events2, provider)
                results[-1] = (op, provider, provider, status2)
            except Exception as e:
                print(f"[Complexification re-test] {op}: {e}")

    # output dirs
    map_name = map_execution_provider(provider)
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    base_output_dir = os.path.join(project_root, f"opset_{ONNX_OPSET_VERSION}", map_name)
    os.makedirs(base_output_dir, exist_ok=True)

    # reports (excel + readme)
    generate_report(results, provider, base_output_dir, models_dir)
    generate_report_aggregated(results, provider, base_output_dir, models_dir)


    # Construire la carte d’inférence : op -> bool(success)
    def _infer_outcome(status: str) -> str:
        s = status.upper()
        if s.startswith("SUCCESS WITH FALLBACK") or "FALLBACK" in s:
            return "SUCCESS_WITH_FALLBACK"
        if s.startswith("SUCCESS"):
            return "SUCCESS"
        return "FAIL"

    inference_outcome_map = {}
    for (op_name, _prov, _used, status) in results:
        outcome = _infer_outcome(status)
        prev = inference_outcome_map.get(op_name)
        if prev is None:
            inference_outcome_map[op_name] = outcome
        else:
            # priorité: SUCCESS > SUCCESS_WITH_FALLBACK > FAIL
            rank = {"SUCCESS": 2, "SUCCESS_WITH_FALLBACK": 1, "FAIL": 0}
            if rank[outcome] > rank[prev]:
                inference_outcome_map[op_name] = outcome
    
    training_status_map = None
    if has_ort_training() and provider in ("CPUExecutionProvider", "CUDAExecutionProvider"):
        training_artifacts_dir = os.path.join(base_output_dir, "training_artifacts")
        _basic, _ms = load_ops(test_file)
        all_ops = _basic + _ms
        _, training_status_map = _run_training_for_provider(
            provider=provider,
            ops_to_test=all_ops,
            models_dir=models_dir,
            base_output_dir=base_output_dir,
            artifacts_root=training_artifacts_dir,
            inference_outcome_map=inference_outcome_map,  # <<< nouveau param
        )


    # README with extra training badges if available
    generate_readme_split(results, provider, base_output_dir, training_status_map=training_status_map, opset_version=ONNX_OPSET_VERSION)


def run_all_providers_and_generate_reports(
    test_file: str = "test.txt",
    base_profiling_dir: str = "profiling",
    base_models_dir: str = "models",
):
    """
    Detects available providers and runs the test suite (inference + optional training) for each.
    """
    providers = ort.get_available_providers()
    if not providers:
        print("No ONNXRuntime execution providers found.")
        return

    for provider in providers:
        profiling_dir = os.path.join(base_profiling_dir, provider)
        models_dir = os.path.join(base_models_dir, provider)
        print(f"\n=== Running tests with provider: {provider} ===")
        run_tests_and_generate_reports(
            provider=provider,
            test_file=test_file,
            profiling_dir=profiling_dir,
            models_dir=models_dir,
        )

# === ADD: exécution ciblée sur une liste d'EP et une liste d'opsets ===
def run_selected_providers_and_generate_reports(
    providers_to_run: list[str],
    test_file: str = "test.txt",
    base_profiling_dir: str = "profiling",
    base_models_dir: str = "models",
    opsets: list[int] = (22,),
):
    """
    Exécute la suite pour une sélection d'EP (noms ORT exacts) et pour N opsets.
    - Sans modifier les builders ops/* : on commute l'opset via configure_opset()
    """
    available = set(ort.get_available_providers())
    to_run = [p for p in providers_to_run if p in available]
    missing = [p for p in providers_to_run if p not in available]
    if missing:
        print(f"[warn] Skipping unavailable providers: {missing}")

    if not to_run:
        print("[info] No matching providers to run in this environment.")
        return

    for opset in opsets:
        print(f"\n=== Configure opset={opset} ===")
        configure_opset(opset)
        for provider in to_run:
            profiling_dir = os.path.join(base_profiling_dir, f"opset_{opset}", provider)
            models_dir = os.path.join(base_models_dir, f"opset_{opset}", provider)
            os.makedirs(profiling_dir, exist_ok=True)
            os.makedirs(models_dir, exist_ok=True)
            print(f"\n=== Running tests with provider: {provider} (opset={opset}) ===")
            run_tests_and_generate_reports(
                provider=provider,
                test_file=test_file,
                profiling_dir=profiling_dir,
                models_dir=models_dir,
            )
